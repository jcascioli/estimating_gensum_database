#NEXT: copy brs file template to prj folder then replace paths in program to point to new location
#NEXT: transfer user input information to gensum
#next: transfer brs first column company to gensum
#next: transfer final brs results to historical database 
#import libraries
import numpy as np
import pandas as pd
import urllib
import os
import openpyxl
#import winshell
from win32com.client import Dispatch
from shutil import copy as cp
import datetime
#from uszipcode import SearchEngine, SimpleZipcode, Zipcode


def import_csv(file):
    #import csv file
    df=pd.read_csv(file)
    return df
def get_download_path():
    #return default download path for linux or windows
        if os.name == 'nt':
            import winreg
            sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
                location = winreg.QueryValueEx(key, downloads_guid)[0]
            return location
        else:
            return os.path.join(os.path.expanduser('~'), 'downloads')


def master_program():

    # GET PATH: users download directory on windows computer
    download_path = get_download_path()
    
    # CREATE PATH: creats folder path of default download path with default folder where gensum file data is to be extracted from is to be placed
    project_setup_directory= download_path +r'\_02-estimating_gensum_database'
    
    # CREATE PATH:Creates folder path with a name of the time program was ran
    time_now = datetime.datetime.now()
    current_time =str(time_now)[0:19].replace(":","-")
    print(current_time)
    prj_folder = project_setup_directory+r"\\"+current_time
    
   
    # CREATE FOLDER: create project folder with name of time program ran
    #  This avoids confusion if ran several times
    #  This is where all files resulting from running program will be stored
    if os.path.exists(prj_folder):
        return
    else:
        os.makedirs(prj_folder)
    
    # CREATE PATH: Location of file that data is extracted from 
    gensum_run_path=project_setup_directory +r'\gensum_run.xlsx'
    
    # OPEN FILE: gensum_run to extract info from
    try:
        wb_exist_gensum=openpyxl.load_workbook(gensum_run_path,data_only=True)
    except:
        print("Error when opening workbook.\nTry deleting all other worksheets in excel document and resaving file.\nMake sure file is in correct location and named 'gensum_run.xlsx'\nIf still does not work contact: Jonathancascioli@gmail.com")
    else:
        print("workbook successfully imported.")

    # TELL USER: what worksheets are in excel document
    print("The following worksheets are in the excel document.")
    
    for sheet in wb_exist_gensum:
       print(sheet.title)
    
    # INPUT: Ask user which worksheet he wants to open
    worksheet_open_question=input("Which worksheet would you like to open?:\n")
    
    # TRY: To open user input worksheet
    try:
        ws = wb_exist_gensum[worksheet_open_question]
    except:
        print("That Worksheet does not exist\n")
    else:
        print("Worksheet successfully opened\n")
    
    #FIND CELL: Find cell with "start" in it to set the beginning of a range to pull data from
    for row in ws.iter_rows():
        for cell in row:
            if cell.value =="START":
                start_loc=cell.coordinate
                start_offset=ws[start_loc].offset(1,0).coordinate
    print("Start Location:"+start_loc+"\n")
    print("Start offset:"+start_offset+"\n")

    #FIND CELL: Find cell with "STOP" in it to set the end of a range to pull data from
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "STOP":
                stop_loc=cell.coordinate
                stop_offset=ws[stop_loc].offset(-1,0).coordinate
    print("Stop location:"+stop_loc+"\n")
    print("Stop Offset:"+stop_offset+"\n")



    # CREATE EXCEL: document to place extracted data into
    new_wb=openpyxl.Workbook()
    
    # CREATE SHEETS: create workbook sheets to place data into
    ws3=new_wb.create_sheet("PROJECT_INFO",0)
    ws4=new_wb.create_sheet("PROJECT_FINANCIALS",1)
    ws2=new_wb.create_sheet("BID_PACKAGES",2)
    

    #SET HEADERS: Set headers for each column
    ws2.cell(row=1,column=1,value="Time Ran")
    ws2.cell(row=1,column=2,value="AOP Number")
    ws2.cell(row=1,column=3,value="Project")
    ws2.cell(row=1,column=4,value="BP#")
    ws2.cell(row=1,column=5,value="Bid Package Description")
    ws2.cell(row=1,column=6,value="Package_Total")
    ws2.cell(row=1,column=7,value="Subcontractor Carried")
    ws2.cell(row=1,column=8,value="Cost Per Square Foot")
    
    ws3.cell(row=1,column=1,value="Time Ran")
    ws3.cell(row=1,column=2,value="Project Name")
    ws3.cell(row=1,column=3,value="Address")
    ws3.cell(row=1,column=4,value="Sector")
    ws3.cell(row=1,column=5,value="CRM Category")
    ws3.cell(row=1,column=6,value="Private/Public")
    ws3.cell(row=1,column=7,value="Proposal Type")
    ws3.cell(row=1,column=8,value="Client")
    ws3.cell(row=1,column=9,value="Architect")
    ws3.cell(row=1,column=10,value="City")
    ws3.cell(row=1,column=11,value="State")
    ws3.cell(row=1,column=12,value="Est Project #")
    ws3.cell(row=1,column=13,value="Estimate Date")
    ws3.cell(row=1,column=14,value="AOP Number")
    ws3.cell(row=1,column=15,value="Total SQ FT")

    ws4.cell(row=1,column=1,value="Time Ran")
    ws4.cell(row=1,column=2,value="Indirect Line Item")
    ws4.cell(row=1,column=3,value="Indirect Totals")
    ws4.cell(row=1,column=4,value="Percentage Applied")
    ws4.cell(row=1,column=5,value="Cost Per Sq Ft")
    ws4.cell(row=1,column=6,value="Project Name")
    ws4.cell(row=1,column=7,value="AOP")

    #SET ROWS: indirect Line Item Rows
    ws4.cell(row=2,column=2,value="Direct Work Total")
    ws4.cell(row=3,column=2,value="SDI")
    ws4.cell(row=4,column=2,value="Bond or Corporate Guarantee")
    ws4.cell(row=5,column=2,value="Insurance(GL & WC)")
    ws4.cell(row=6,column=2,value="Insurance (OCP&L)")
    ws4.cell(row=7,column=2,value="Builders Risk")
    ws4.cell(row=8,column=2,value="General Conditions(W/O insurance)")
    ws4.cell(row=9,column=2,value="Building Permit")
    ws4.cell(row=10,column=2,value="Fee")
    ws4.cell(row=11,column=2,value="Precon")
    ws4.cell(row=12,column=2,value="Escalation")
    ws4.cell(row=13,column=2,value="Contigency")
    ws4.cell(row=14,column=2,value="Additional Contigency")
    ws4.cell(row=15,column=2,value="Tax")
    ws4.cell(row=16,column=2,value="Total After Indirect Costs")
    ws4.cell(row=17,column=2,value="General Add/Deduct")
    ws4.cell(row=18 ,column=2,value="Final Total")
            
   
    counter=2
    # PULL RANGE: pull Bid Division
    division_cell_range=ws[start_offset:stop_offset]
    for cell, in division_cell_range:
        division=cell.value
        print(division)
        ws2.cell(row=counter,column=4,value=division)
        counter+=1
            
    #PULL RANGE: pull bid package description     
    counter=2
    bid_package_start=ws[start_offset].offset(0,1).coordinate
    bid_package_stop=ws[stop_offset].offset(0,1).coordinate
    bid_package_cell_range=ws[bid_package_start:bid_package_stop]
    for cell, in bid_package_cell_range:
        bid_package=cell.value
        print(bid_package)
        ws2.cell(row=counter,column=5,value=bid_package)
        counter+=1


    #PULL RANGE:  pull total cost column
    counter=2
    total_col_start=ws[start_offset].offset(0,2).coordinate
    total_col_stop=ws[stop_offset].offset(0,2).coordinate
    total_col_range=ws[total_col_start:total_col_stop]
    for cell, in total_col_range:
        tol_value=cell.value
        print(tol_value)
        ws2.cell(row=counter,column=6,value=tol_value)
        counter+=1

    #FIND CELL: Find cell with "Subcontractor Carried" in it to set the end of a range to pull data from
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Subcontractor Carried":
                sub_carried_loc=cell.coordinate
                sub_carried_offset=ws[sub_carried_loc].offset(2,0).coordinate
    print("Start location:"+sub_carried_loc+"\n")
    print("Start Offset:"+sub_carried_offset+"\n\n\n")
    
    #FIND CELL: Find cell with "STOP 2" in it to set the end of a range to pull data from
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "STOP_2":
                stop2_loc=cell.coordinate
                stop2_offset=ws[stop2_loc].offset(-1,0).coordinate
    print("Stop location:"+stop2_loc+"\n")
    print("Stop Offset:"+stop2_offset+"\n\n\n")


    #PULL RANGE: Pull subcontractor carried column
    counter=2
    sub_carried_range=ws[sub_carried_offset:stop2_offset]
    for cell, in sub_carried_range:
        sub_value=cell.value
        print(sub_value)
        ws2.cell(row=counter,column=7,value=sub_value)
        counter+=1

    #FIND CELL: indirect cost per sq ft column
    for row in ws.iter_rows():
        for cell in row:
            if cell.value =="direct cost per sq ft":
                coordinate=cell.coordinate
                print("Direct Cost Per Sq Ft found at:  "+coordinate+"\n\n")
                idcsf_loc=ws[coordinate].offset(2,0).coordinate

    #FIND CELL: STOP_3
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="STOP_3":
                coordinate=cell.coordinate
                print("STOP_3 found at:  "+coordinate+"\n\n")
                stop_3_loc=ws[coordinate].offset(-1,0).coordinate

    #PULL RANGE & ITTERATE: inderect cost per sq ft column
    x=2
    idcsf_range=ws[idcsf_loc:stop_3_loc]
    for cell, in idcsf_range:
        idcsf_value=cell.value
        print(idcsf_value)
        ws2.cell(row=x,column=8,value=idcsf_value)
        x+=1
    #FIND CELL: FIND Project Name Cell, then offset over 1 column to retrieve project name info
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Project Name:":
                prj_n_coord=cell.coordinate
                print(prj_n_coord)
                prj_n_value=ws[prj_n_coord].offset(0,1).value
                prj_n_info_coord=ws[prj_n_coord].offset(0,1).coordinate
    print("Project Name is:"+prj_n_value+"\n\n\n")

    #FIND CELL: AOP number
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Aop:":
                aop_coord=cell.coordinate
                print(aop_coord)
                aop_value=ws[aop_coord].offset(0,1).value
                aop_info_coord=ws[aop_coord].offset(0,1).coordinate
                print(aop_info_coord)

    #PULL RANGE: Project information
    counter=2
    prj_info_range=ws[prj_n_info_coord:aop_info_coord]
    for cell, in prj_info_range:
        prj_info_value=cell.value
        print(prj_info_value)
        ws3.cell(row=2,column=counter,value=prj_info_value)
        counter+=1

    #FIND GSF:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Total SQ FT":
                sf_coord=cell.coordinate
                print("Total SQ Ft Found at:  "+sf_coord+"\n\n")
                sf_value=ws[sf_coord].offset(0,1).value
                ws3.cell(row=2,column=15,value=sf_value)

    #INSERT DATE: Insert current time into individual worksheets
    ws3.cell(row=2,column=1,value=current_time)

    #ITTERATE DATE: itterate current time(got earlier) over every row in first column to ref when run
    counter=2
    for x in range(len(total_col_range)):
        ws2.cell(row=counter,column=1,value=current_time)
        counter+=1
    
    #ITTERATE DATE:  over indirect worksheet
    x=1
    y=2
    while x<18:
        ws4.cell(row=y,column=1,value=current_time)
        x+=1
        y+=1
            

    #PULL RANGE: Pull project name and apply over every row of bid package length
    counter=2
    for x in range(len(total_col_range)):
        ws2.cell(row=counter,column=3,value=prj_n_value)
        counter+=1
    
    #ITTERATE PROJECT NAME: Project name over every row of Project financials page
    l=2
    x=2
    while l<19:
        ws4.cell(row=x,column=6,value=prj_n_value)
        x+=1
        l+=1
        
    #PULL RANGE: Pull aop and apply over every row of bid package length
    counter=2
    for x in range(len(total_col_range)):
        ws2.cell(row=counter,column=2,value=aop_value)
        counter+=1

    #ITTERATE AOP: over every row of project financials page
    l=2
    x=2
    while l<19:
        ws4.cell(row=x,column=7,value=aop_value)
        x+=1
        l+=1
    ## PROJECT FINANCIALS/ INDIRECTS PAGE ##
   
    row_set=2
    #FIND CELL: DIRECT WORK
    for row in ws.iter_rows():
        for cell in row:
            if cell.value =="Direct Work Total (NO SDI or Bonds)":
                dw_coord=cell.coordinate
                print("Direct Work Found at:  "+dw_coord+"\n\n")
                dw_value=ws[dw_coord].offset(0,1).value
                ws4.cell(row=row_set,column=3,value=dw_value)
                #Percentage Applied
                dw_value=ws[dw_coord].offset(0,2).value
                ws4.cell(row=row_set,column=4,value=dw_value)
                #Cost Per SQ Ft
                dw_value=ws[dw_coord].offset(0,3).value
                ws4.cell(row=row_set,column=5,value=dw_value)

    row_set+=1
    #FIND CELL: SDI
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="SDI":
                sdi_coord=cell.coordinate
                print("SDI Found at:  "+sdi_coord+"\n\n")
                sdi_value=ws[sdi_coord].offset(0,1).value
                ws4.cell(row=row_set,column=3,value=sdi_value)
                #Percentage Applied
                sdi_value=ws[sdi_coord].offset(0,2).value
                ws4.cell(row=row_set,column=4,value=sdi_value)
                #Cost Per SQ Ft
                sdi_value=ws[sdi_coord].offset(0,3).value
                ws4.cell(row=row_set,column=5,value=sdi_value)

    row_set+=1
    #FIND CELL: Bonds or Corporate Guarantee
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Bonds or Corporate Guarantee":
                bcg_coord=cell.coordinate
                print("Bond and Corp Guarantee found at:  "+bcg_coord+"\n\n")
                bcg_value=ws[bcg_coord].offset(0,1).value
                ws4.cell(row=row_set,column=3,value=bcg_value)
                #Percentage Applied
                bcg_value=ws[bcg_coord].offset(0,2).value
                ws4.cell(row=row_set,column=4,value=bcg_value)
                #Cost per Sq ft
                bcg_value=ws[bcg_coord].offset(0,3).value
                ws4.cell(row=row_set,column=5,value=bcg_value)
    
    row_set+=1
    #FIND CELL: Insurance GL&WC
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Insurance GL and WC":
                ins_gl_coord=cell.coordinate
                print("Insurance GL & WC found at:  "+ins_gl_coord+"\n\n")
                counter_loop=1
                x=1
                y=3
                while counter_loop<4:
                    ins_gl_value=ws[ins_gl_coord].offset(0,x).value
                    print(ins_gl_value)
                    ws4.cell(row=row_set,column=y,value=ins_gl_value)
                    x+=1
                    y+=1
                    counter_loop+=1

    row_set+=1                               
    #FIND CELL: Insurance OCPL
    for row in ws.iter_rows():
         for cell in row:
             if cell.value=="Insurance OCPL":
                 ins_oc_coord=cell.coordinate
                 print("Insurance OCPL found at:  "+ins_oc_coord+"\n\n")
                 counter_loop=1
                 x=1
                 y=3
                 while counter_loop<4:
                     ins_oc_value=ws[ins_oc_coord].offset(0,x).value
                     ws4.cell(row=row_set,column=y,value=ins_oc_value)
                     x+=1
                     y+=1
                     counter_loop+=1
    row_set+=1
    #FIND CELL: Builders Risk
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Builders Risk":
                br_coord=cell.coordinate
                print("Builders Risk Found at  "+br_coord+"\n\n")
                counter_loop=1
                x=1
                y=3
                while counter_loop<4:
                    br_value=ws[br_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=br_value)
                    x+=1
                    y+=1
                    counter_loop+=1

    row_set+=1
    #FIND CELL: General Conditions w/o insurannce
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="General Conditions (w/o Insurance)":
                gc_coord=cell.coordinate
                print("General Conditions (w/o insurance) found at:  "+gc_coord+"\n\n")
                counter_loop=1
                x=1
                y=3
                while counter_loop<4:
                    gc_value=ws[gc_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=gc_value)
                    x+=1
                    y+=1
                    counter_loop+=1
    row_set+=1
    #FIND CELL: Building Permit
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Building Permit":
                bp_coord=cell.coordinate
                print("Building Permit Found at:  "+bp_coord+"\n\n")
                cl=1
                x=1
                y=3
                while cl<4:
                    bp_value=ws[bp_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=bp_value)
                    x+=1
                    y+=1
                    cl+=1
    row_set+=1
    #FIND CELL: FEE
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Fee":
                fee_coord=cell.coordinate
                print("Fee found at:  "+fee_coord+"\n\n")
                l=1
                x=1
                y=3
                while l<4:
                    fee_value=ws[fee_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=fee_value)
                    x+=1
                    y+=1
                    l+=1
    row_set+=1
    #FIND CELL: Precon
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Precon":
                precon_coord=cell.coordinate
                print("Precon found at:  "+precon_coord+"\n\n")
                x=1
                y=3
                l=1
                while l<4:
                    precon_value=ws[precon_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=precon_value)
                    x+=1
                    y+=1
                    l+=1
    row_set+=1                    
    #FIND CELL: Escalation
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Indirect Escalation":
                escalation_coord=cell.coordinate
                print("Indirect Escalation found at:  "+escalation_coord+"\n\n")
                l=1
                x=1
                y=3
                while l<4:
                    escalation_value=ws[escalation_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=escalation_value)
                    x+=1
                    y+=1
                    l+=1
    row_set+=1
    #FIND CELL: Contigency
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Contingency":
                contingency_coord=cell.coordinate
                print("Contingency found at:  "+contingency_coord+"\n\n")
                l=1
                x=1
                y=3
                while l<4:
                    contingency_value=ws[contingency_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=contingency_value)
                    x+=1
                    y+=1
                    l+=1
    row_set+=1
    #FIND CELL: additional contigency
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Additional Contingency":
                ad_cont_coord=cell.coordinate
                print("Additional Contingency  found at:  "+ad_cont_coord+"\n\n")
                l=1
                y=3
                x=1
                while l<4:
                    ad_cont_value=ws[ad_cont_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=ad_cont_value)
                    l+=1
                    y+=1
                    x+=1
    row_set+=1
    #FIND CELL: TAX
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Tax":
                tax_coord=cell.coordinate
                print("Tax found at:  "+tax_coord+"\n\n")
                l=1
                y=3
                x=1
                while l<4:
                    tax_value=ws[tax_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=tax_value)
                    l+=1
                    y+=1
                    x+=1
    row_set+=1
    #FIND CELL: Total After indirect Costs
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="Total After Indirect Costs":
                ta_coord=cell.coordinate
                print("Total After Indirect Costs found at:  "+ta_coord+"\n\n")
                l=1
                y=3
                x=1
                while l<4:
                    ta_value=ws[ta_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=ta_value)
                    l+=1
                    y+=1
                    x+=1
                
    row_set+=1
    #FIND Cell: General Add/Deduct
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="General Add/Deduct":
                gad_coord=cell.coordinate
                print("General Add/Deduct found at:  "+gad_coord+"\n\n")
                l=1
                y=3
                x=1
                while l<4:
                    gad_value=ws[gad_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=gad_value)
                    l+=1
                    y+=1
                    x+=1
    row_set+=1
    #FIND CELL: Final Total
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="FINAL TOTAL":
                ft_coord=cell.coordinate
                print("Final Total found at:  "+ft_coord+"\n\n")
                l=1
                y=3
                x=1
                while l<4:
                    ft_value=ws[ft_coord].offset(0,x).value
                    ws4.cell(row=row_set,column=y,value=ft_value)
                    x+=1
                    y+=1
                    l+=1


    
    # CREATE PATH: path to use for resulting file of program
    gensum_prj_file = prj_folder+ r"\\"+current_time+"__-__gensum.xlsx"

    # SAVE EXCEL: save excel file to path location
    new_wb.save(gensum_prj_file)






master_program()
